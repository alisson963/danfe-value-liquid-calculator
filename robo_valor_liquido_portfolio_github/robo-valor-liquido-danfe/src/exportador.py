from pathlib import Path
from datetime import datetime
import os
import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from utils import converter_numero
COLUNAS_RESULTADO_FINAL = ['documento_id', 'arquivo', 'chave_acesso', 'numero_nfe', 'serie_nfe', 'data_emissao', 'emitente', 'destinatario', 'valor_total_nota_documento', 'numero_item', 'codigo_produto', 'descricao', 'ncm', 'cst', 'cfop', 'unidade', 'quantidade', 'valor_unitario', 'valor_total', 'base_icms_documento', 'valor_icms_documento', 'valor_ipi_documento', 'aliquota_icms', 'aliquota_ipi', 'pis_percentual_usado', 'cofins_percentual_usado', 'valor_ipi_calculado', 'valor_icms_calculado', 'base_pis_cofins', 'valor_pis', 'valor_cofins', 'valor_liquido_total', 'valor_liquido_unitario']
COLUNAS_DOCUMENTOS = ['documento_id', 'arquivo', 'status', 'itens_encontrados', 'chave_acesso', 'numero_nfe', 'serie_nfe', 'data_emissao', 'emitente', 'destinatario', 'valor_total_nota_documento', 'soma_valor_total_itens', 'soma_valor_liquido_total', 'detalhe']
COLUNAS_NUMERICAS = ['quantidade', 'valor_unitario', 'valor_total', 'base_icms_documento', 'valor_icms_documento', 'valor_ipi_documento', 'aliquota_icms', 'aliquota_ipi', 'pis_percentual_usado', 'cofins_percentual_usado', 'valor_ipi_calculado', 'valor_icms_calculado', 'base_pis_cofins', 'valor_pis', 'valor_cofins', 'valor_liquido_total', 'valor_liquido_unitario', 'valor_total_nota_documento', 'soma_valor_total_itens', 'soma_valor_liquido_total']
MOEDA_COLS = ['valor_unitario', 'valor_total', 'base_icms_documento', 'valor_icms_documento', 'valor_ipi_documento', 'valor_total_nota_documento', 'soma_valor_total_itens', 'valor_ipi_calculado', 'valor_icms_calculado', 'base_pis_cofins', 'valor_pis', 'valor_cofins', 'valor_liquido_total', 'soma_valor_liquido_total', 'valor_liquido_unitario']
PERCENTUAL_COLS = ['aliquota_icms', 'aliquota_ipi', 'pis_percentual_usado', 'cofins_percentual_usado']
NOMES_COLUNAS = {'documento_id': 'Documento / Identificação', 'arquivo': 'Arquivo', 'tipo_arquivo': 'Tipo Arquivo', 'chave_acesso': 'Chave de Acesso', 'numero_nfe': 'Nº NF-e', 'serie_nfe': 'Série', 'data_emissao': 'Data Emissão', 'emitente': 'Emitente', 'destinatario': 'Destinatário', 'numero_item': 'Item', 'codigo_produto': 'Código Produto', 'descricao': 'Descrição do Produto / Serviço', 'ncm': 'NCM/SH', 'cst': 'O/CST', 'cfop': 'CFOP', 'unidade': 'UN', 'quantidade': 'Quantidade', 'valor_unitario': 'Valor Unitário', 'valor_total': 'Valor Total', 'base_icms_documento': 'B. Cálc. ICMS Documento', 'valor_icms_documento': 'Valor ICMS Documento', 'valor_ipi_documento': 'Valor IPI Documento', 'aliquota_icms': 'Alíq. ICMS %', 'aliquota_ipi': 'Alíq. IPI %', 'pis_percentual_usado': 'PIS % Usado', 'cofins_percentual_usado': 'COFINS % Usado', 'valor_ipi_calculado': 'Valor IPI Calculado', 'valor_icms_calculado': 'Valor ICMS Calculado', 'base_pis_cofins': 'Base PIS/COFINS', 'valor_pis': 'Valor PIS', 'valor_cofins': 'Valor COFINS', 'valor_liquido_total': 'Valor Líquido Total', 'valor_liquido_unitario': 'Vl Lq Unitário', 'status': 'Status', 'itens_encontrados': 'Itens Encontrados', 'valor_total_nota_documento': 'Valor Total da Nota', 'soma_valor_total_itens': 'Soma Valor Total Itens', 'soma_valor_liquido_total': 'Soma Valor Líquido Total', 'detalhe': 'Detalhe', 'erro': 'Erro'}

def _renomear_colunas(df: pd.DataFrame) -> pd.DataFrame:
    return df.rename(columns={col: NOMES_COLUNAS.get(col, col) for col in df.columns})

def _garantir_colunas(df: pd.DataFrame, colunas: list[str]) -> pd.DataFrame:
    for coluna in colunas:
        if coluna not in df.columns:
            df[coluna] = ''
    return df[colunas]

def preparar_dataframe(resultados: list[dict]) -> pd.DataFrame:
    df = pd.DataFrame(resultados or [])
    df = _garantir_colunas(df, COLUNAS_RESULTADO_FINAL)
    return converter_colunas_numericas(df)

def preparar_dataframe_erros(erros: list[dict]) -> pd.DataFrame:
    colunas = ['arquivo', 'erro']
    df = pd.DataFrame(erros or [])
    return _garantir_colunas(df, colunas)

def preparar_dataframe_documentos(documentos: list[dict], df_itens: pd.DataFrame) -> pd.DataFrame:
    df_docs = pd.DataFrame(documentos or [])
    df_docs = _garantir_colunas(df_docs, COLUNAS_DOCUMENTOS)
    if not df_itens.empty:
        agrupado = df_itens.groupby('arquivo', dropna=False).agg(chave_acesso=('chave_acesso', 'first'), numero_nfe=('numero_nfe', 'first'), serie_nfe=('serie_nfe', 'first'), data_emissao=('data_emissao', 'first'), emitente=('emitente', 'first'), destinatario=('destinatario', 'first'), valor_total_nota_documento=('valor_total_nota_documento', 'first'), soma_valor_total_itens=('valor_total', 'sum'), soma_valor_liquido_total=('valor_liquido_total', 'sum'), documento_id=('documento_id', 'first')).reset_index()
        df_docs = df_docs.drop(columns=['documento_id', 'chave_acesso', 'numero_nfe', 'serie_nfe', 'data_emissao', 'emitente', 'destinatario', 'valor_total_nota_documento', 'soma_valor_total_itens', 'soma_valor_liquido_total'], errors='ignore')
        df_docs = df_docs.merge(agrupado, on='arquivo', how='left')
    df_docs = _garantir_colunas(df_docs, COLUNAS_DOCUMENTOS)
    return converter_colunas_numericas(df_docs)

def converter_colunas_numericas(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    for coluna in COLUNAS_NUMERICAS:
        if coluna in df.columns:
            df[coluna] = df[coluna].apply(lambda x: float(converter_numero(x)) if str(x).strip() not in ('', 'nan', 'None') else 0)
    return df

def gerar_caminho_com_data(caminho_saida: Path) -> Path:
    agora = datetime.now().strftime('%Y%m%d_%H%M%S')
    return caminho_saida.with_name(f'{caminho_saida.stem}_{agora}{caminho_saida.suffix}')

def _criar_resumo(df: pd.DataFrame, df_docs: pd.DataFrame) -> pd.DataFrame:
    arquivos_com_sucesso = int((df_docs['status'] == 'OK').sum()) if not df_docs.empty else 0
    arquivos_com_erro = int((df_docs['status'] != 'OK').sum()) if not df_docs.empty else 0
    return pd.DataFrame({'Indicador': ['Arquivos encontrados', 'Arquivos processados com sucesso', 'Arquivos com erro/atenção', 'Itens encontrados', 'Soma valor total', 'Soma valor líquido total', 'Observação'], 'Valor': [len(df_docs), arquivos_com_sucesso, arquivos_com_erro, len(df), float(df['valor_total'].sum()) if not df.empty else 0, float(df['valor_liquido_total'].sum()) if not df.empty else 0, 'Use a aba Resultado Final para conferir chave, arquivo, NF, item e cálculo linha por linha.']})

def salvar_excel(resultados: list[dict], caminho_saida: Path, erros: list[dict] | None=None, documentos: list[dict] | None=None) -> Path:
    caminho_saida = Path(caminho_saida)
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)
    df = preparar_dataframe(resultados)
    df_docs = preparar_dataframe_documentos(documentos or [], df)
    df_erros = preparar_dataframe_erros(erros or [])
    resumo = _criar_resumo(df, df_docs)
    temporario = caminho_saida.with_name(f"~tmp_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}_{caminho_saida.name}")
    with pd.ExcelWriter(temporario, engine='openpyxl') as writer:
        _renomear_colunas(df).to_excel(writer, sheet_name='Resultado Final', index=False)
        _renomear_colunas(df_docs).to_excel(writer, sheet_name='Documentos', index=False)
        resumo.to_excel(writer, sheet_name='Resumo', index=False)
        _renomear_colunas(df_erros).to_excel(writer, sheet_name='Erros', index=False)
        wb = writer.book
        for ws in wb.worksheets:
            aplicar_formatacao(ws)
            adicionar_tabela(ws)
        if 'Resultado Final' in wb.sheetnames:
            ws = wb['Resultado Final']
            ws.freeze_panes = 'A2'
            aplicar_formatos_numericos(ws)
            destacar_colunas_principais(ws)
        if 'Documentos' in wb.sheetnames:
            ws_docs = wb['Documentos']
            ws_docs.freeze_panes = 'A2'
            aplicar_formatos_numericos(ws_docs)
        if 'Resumo' in wb.sheetnames:
            ws_resumo = wb['Resumo']
            for row in range(2, ws_resumo.max_row + 1):
                if row in (6, 7):
                    ws_resumo[f'B{row}'].number_format = 'R$ #,##0.00'
            ws_resumo.column_dimensions['A'].width = 36
            ws_resumo.column_dimensions['B'].width = 90
    try:
        os.replace(temporario, caminho_saida)
        return caminho_saida
    except PermissionError:
        caminho_alternativo = gerar_caminho_com_data(caminho_saida)
        os.replace(temporario, caminho_alternativo)
        return caminho_alternativo
    except OSError:
        caminho_alternativo = gerar_caminho_com_data(caminho_saida)
        os.replace(temporario, caminho_alternativo)
        return caminho_alternativo

def aplicar_formatos_numericos(ws):
    headers = {cell.value: cell.column for cell in ws[1]}
    for nome in [NOMES_COLUNAS.get(c, c) for c in MOEDA_COLS]:
        col = headers.get(nome)
        if col:
            for col_cells in ws.iter_cols(min_col=col, max_col=col, min_row=2):
                for c in col_cells:
                    c.number_format = 'R$ #,##0.00'
    for nome in [NOMES_COLUNAS.get(c, c) for c in PERCENTUAL_COLS]:
        col = headers.get(nome)
        if col:
            for col_cells in ws.iter_cols(min_col=col, max_col=col, min_row=2):
                for c in col_cells:
                    c.number_format = '0.00'
    quantidade_col = headers.get('Quantidade')
    if quantidade_col:
        for col_cells in ws.iter_cols(min_col=quantidade_col, max_col=quantidade_col, min_row=2):
            for c in col_cells:
                c.number_format = '#,##0.0000'

def aplicar_formatacao(ws):
    header_fill = PatternFill('solid', fgColor='1F4E78')
    header_font = Font(color='FFFFFF', bold=True)
    thin = Side(style='thin', color='D9E2F3')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=False)
    largura_fixa = {'Documento / Identificação': 44, 'Arquivo': 38, 'Chave de Acesso': 48, 'Emitente': 36, 'Destinatário': 36, 'Descrição do Produto / Serviço': 48, 'Detalhe': 55, 'Erro': 55}
    for column_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(column_cells[0].column)
        header = column_cells[0].value
        for cell in column_cells:
            value = '' if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        largura = largura_fixa.get(header, min(max(max_len + 2, 12), 32))
        ws.column_dimensions[col_letter].width = largura
    ws.row_dimensions[1].height = 34

def adicionar_tabela(ws):
    if ws.max_row < 1 or ws.max_column < 1:
        return
    ref = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'
    nome = 'Tabela_' + reformatar_nome_tabela(ws.title)
    try:
        tab = Table(displayName=nome, ref=ref)
        style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        ws.add_table(tab)
    except Exception:
        pass

def reformatar_nome_tabela(nome: str) -> str:
    limpo = ''.join((ch if ch.isalnum() else '_' for ch in nome))
    return limpo[:20] or 'Dados'

def destacar_colunas_principais(ws):
    headers = {cell.value: cell.column for cell in ws[1]}
    fill = PatternFill('solid', fgColor='D9EAF7')
    principais = ['Documento / Identificação', 'Arquivo', 'Chave de Acesso', 'Nº NF-e', 'Código Produto', 'Valor Total', 'Valor Líquido Total', 'Vl Lq Unitário']
    for nome in principais:
        col = headers.get(nome)
        if not col:
            continue
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=col).fill = fill if row != 1 else PatternFill('solid', fgColor='1F4E78')
